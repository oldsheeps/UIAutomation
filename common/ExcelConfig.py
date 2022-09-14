import os

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

PASS_VALUE = 'Pass'         # 单元格文本-Pass
FAILED_VALUE = 'Failed'     # 单元格文本-Failed
CONTINUE_VALUE = 'Continue' # 单元格文本-Continue
HYPERLINK_VALUE = 'Picture' # 超链接文本-Screenshot
CELL_STYLE_RED = 'FF0000'   # 单元格样式-红色填充
CELL_STYLE_GREEN = '00FF00' # 单元格样式-绿色填充
CELL_STYLE_GRAY = 'E8E8E8'  # 单元格样式-灰色填充
CELL_STYLE_BLUE = '0000FF'  # 单元格样式-灰色填充
CELL_STYLE_FONT = 'solid'   # 单元格样式-字体加粗
DEFAULT_ROW_HEIGHT = 30     # 单元格样式-默认行高

class Excel(object):

    def __init__(self, file_name, file_rebuild=False):
        """
        Excel操作
        :param file_name:    文件路径
        :param file_rebuild: 是否重建
        """
        self.file_name = file_name
        self.file_rebuild = file_rebuild
        self.excel = self.get_excel()


    def close(self):
        """关闭excel文件"""
        self.excel.close()


    def save(self, file_name):
        """保存excel文件"""
        self.excel.save(file_name)


    def get_excel(self):
        """ 获取excel对象，没有则自动创建 """
        # 如果文件已存在但是选择重建则删除文件
        if self.file_rebuild and os.path.exists(self.file_name):
            os.remove(self.file_name)
        # 如果文件不存在创建文件
        if not os.path.exists(self.file_name):
            Workbook().save(self.file_name)
        return openpyxl.load_workbook(self.file_name)


    def get_all_sheet_name(self):
        """ 获取所有sheet页名称 """
        return self.excel.sheetnames


    def get_sheet_object(self, sheet_name):
        """
        获取指定sheet页对象，没有则自动创建
        :param sheet_obj: sheet页名称
        :return: sheet页对象
        """
        # 如果sheet页不存在当前excel文件中则新建
        if sheet_name not in self.excel.sheetnames:
            self.excel.create_sheet(sheet_name, 0)
        return self.excel[sheet_name]


    def get_sheet_cols(self, sheet_obj):
        """
        获取指定sheet页中所有列的第一行内容（表头）
        :param sheet_obj: sheet页对象
        :return: 以字典形式返回，cols_name：以列表形式存储所有列数据，cols_num：列的数量
        """
        global index
        # 记录标题字段
        title_field_list = []
        # 使用enumerate有序遍历第一行的每一个字段
        for index, column in enumerate(list(sheet_obj.rows)[0]):
            # 将当前字段添加到列表
            if column.value:
                title_field_list.append(column.value)
            else:
                break
        return {'cols_name': title_field_list, 'cols_num': index + 1}


    def get_sheet_rows(self, sheet_obj):
        """
        获取指定sheet页中所有行数据
        :param sheet_name: sheet页名称
        :return: 以字典形式返回，rows_data：以二维列表形式存储所有行数据，rows_coord：包含所有已使用单元格坐标，rows_num：除去表头总行数
        """
        global index  # 记录除表头外总行数
        rows_data = []  # 记录每行数据
        rows_data_coord = []  # 记录数据坐标
        # 使用enumerate有序遍历所有行
        for index, row in enumerate(sheet_obj.rows):
            # 第一行数据为表头，跳过
            if index == 0:
                continue
            # 当前行的第一个字段有值，则将记录当前行所有列数据和坐标
            if row[0].value:
                rows_data.append([cell.value for cell in row])
                rows_data_coord.append(row)
            else:
                break
        return {'rows_data': rows_data, 'rows_coord': rows_data_coord, 'rows_num': index}


    def get_sheet_data(self, sheet_name):
        """
        获取指定sheet页中所有行、所有列的数据
        :param sheet_name: sheet页名称
        :return: 以列表套字典形式返回，每一个字典都是以表头和行数据组合成的Key-Value
        """
        # 获取sheet对象
        sheet = self.get_sheet_object(sheet_name)
        # 获取sheet页中所有列的表头和列数
        cols = self.get_sheet_cols(sheet)
        cols_name, cols_num = cols['cols_name'], cols['cols_num']
        # 获取sheet页中所有行的坐标和总行数
        rows = self.get_sheet_rows(sheet)
        rows_coord, rows_num = rows['rows_coord'], rows['rows_num']
        # 压缩并转为字典：使用二维循环推导式（先看最右侧循环遍历每一行所使用的单元格，其次再在循环列的数量取出每一个单元格的值）
        return [dict(zip(cols_name, [row[i].value for i in range(cols_num)])) for row in rows_coord]


    def set_column_width(self, sheet_obj):
        """
        宽度自适应：将指定sheet页中所有列设置自适应宽度
        :param sheet_obj: sheet页对象
        :return: 作用于工作簿中指定Sheet页的所有列数据
        """
        # 获取当前sheet页的数据shape
        max_row, max_column = sheet_obj.max_row, sheet_obj.max_column
        for col in range(max_column):
            # 根据列的数字返回对应的字母，暂定每列宽度为1
            max_widths, col_name = 1, get_column_letter(col + 1)
            # 获取当前列下的所有数据，取单元格值内数据最大长度
            for row in range(max_row):
                content_length = len(str(sheet[f'{col_name}{row + 1}'].value or '' or None))
                # 如果单元格内最大长度比假设大则修改
                max_widths = content_length if content_length > max_widths else max_widths
            # 把当前列所有单元格的长度设为最大长度的2倍
            sheet.column_dimensions[col_name].width = max_widths * 2
        self.save(self.file_name)


    def set_row_height(self, sheet_obj, row_height):
        """
       将指定sheet页中所有数据行设置高度
       :param sheet_obj:  sheet页对象
       :param row_height: 高度
       :return: 作用于工作簿中指定Sheet页的所有行数据
       """
        max_row, max_col = sheet_obj.max_row, sheet_obj.max_column
        # 循环设置每一行的高度
        for row in range(1, max_row + 1):
            sheet_obj.row_dimensions[row].height = row_height


    def set_row_style(self, sheet_obj):
        """
        隔行变色：将指定sheet页中所有偶数行设置背景颜色
        :param sheet_obj:  sheet页对象
        :return: 作用于工作簿中指定Sheet页的偶数行
        """
        max_row, max_col = sheet_obj.max_row, sheet_obj.max_column
        color = PatternFill(patternType=CELL_STYLE_FONT, fgColor=CELL_STYLE_GRAY)
        for row in range(1, max_row):
            # 如果i能被2整除就可以得到偶数行
            if row % 2 == 0:
                # 循环设置每一行的背景颜色
                for col in range(1, max_col + 1):
                    sheet.cell(row=row, column=col).fill = color


    def set_cell_style(self, sheet_obj, sheet_row, sheet_col, fgColor, row_height=DEFAULT_ROW_HEIGHT):
        """
       在指定行列的单元格上设置背景背景、加粗字体样式
       :param sheet_obj:  sheet页对象
       :param sheet_row:  sheet页中指定行
       :param sheet_col:  sheet页中指定列
       :param fgColor:    背景颜色
       :return: 作用于工作簿中指定Sheet页的单元格
       """
        sheet_obj.row_dimensions[sheet_row].height = row_height
        sheet_obj.cell(row=sheet_row, column=sheet_col).fill = PatternFill(patternType=CELL_STYLE_FONT, fgColor=fgColor)
        sheet_obj.cell(row=sheet_row, column=sheet_col).font = Font(bold=True)


    def set_cell_value(self, sheet_obj, sheet_row, sheet_col, cell_value, fgColor=CELL_STYLE_GREEN):
        """
       在指定行列的单元格上设置背景颜色、加粗字体样式写入文本值
       :param sheet_obj:  sheet页对象
       :param sheet_row:  sheet页中指定行
       :param sheet_col:  sheet页中指定列
       :param cell_value: 设置文本值
       :param fgColor:    背景颜色
       :return: 作用于工作簿中指定Sheet页的单元格
       """
        sheet_obj.cell(row=sheet_row, column=sheet_col).value = cell_value
        self.set_cell_style(sheet_obj, sheet_row, sheet_col, fgColor)


    def pass_(self, sheet_obj, sheet_row, sheet_col):
        """
       在指定行列的单元格上以绿色背景、加粗字体样式写入Pass文本
       :param sheet_obj: sheet页对象
       :param sheet_row: sheet页中指定行
       :param sheet_col: sheet页中指定列
       :return: 作用于工作簿中指定Sheet页的单元格
       """
        sheet_obj.cell(row=sheet_row, column=sheet_col).value = PASS_VALUE
        self.set_cell_style(sheet_obj, sheet_row, sheet_col, CELL_STYLE_GREEN)


    def failed_(self, sheet_obj, sheet_row, sheet_col):
        """
        在指定行列的单元格上以红色背景、加粗字体样式写入Failed文本
        :param sheet_obj: sheet页对象
        :param sheet_row: sheet页中指定行
        :param sheet_col: sheet页中指定列
        :return: 作用于工作簿中指定Sheet页的单元格
        """
        sheet_obj.cell(row=sheet_row, column=sheet_col).value = FAILED_VALUE
        self.set_cell_style(sheet_obj, sheet_row, sheet_col, CELL_STYLE_RED)


    def continue_(self, sheet_obj, sheet_row, sheet_col):
        """
       在指定行列的单元格上以绿色背景、加粗字体样式写入Continue文本
       :param sheet_obj: sheet页对象
       :param sheet_row: sheet页中指定行
       :param sheet_col: sheet页中指定列
       :return: 作用于工作簿中指定Sheet页的单元格
       """
        sheet_obj.cell(row=sheet_row, column=sheet_col).value = CONTINUE_VALUE
        self.set_cell_style(sheet_obj, sheet_row, sheet_col, CELL_STYLE_GREEN)


    def hyperlink_(self, sheet_obj, sheet_row, sheet_col, file_path, link_name=HYPERLINK_VALUE,
                   fgColor=CELL_STYLE_GREEN):
        """
        在指定行列的单元格上以绿色背景、加粗字体样式写入指定文本超链接
        :param sheet_obj: sheet页对象
        :param sheet_row: sheet页中指定行
        :param sheet_col: sheet页中指定列
        :param file_path: 超链接文件路径
        :param link_name: 超链接显示文本
        :param fgColor:   背景颜色
        :return: 作用于工作簿中指定Sheet页的单元格
        """
        sheet_obj.cell(row=sheet_row, column=sheet_col).value = f'=HYPERLINK("{file_path}","{link_name}")'
        self.set_cell_style(sheet_obj, sheet_row, sheet_col, fgColor)



if __name__ == '__main__':
    path = '../config/关键字框架测试用例Demo-循环.xlsx'
    excel = Excel(path)
    sheets = excel.get_all_sheet_name()
    print("获取所有sheet页：", sheets)
    sheet = excel.get_sheet_object('百度搜索')
    print("get_sheet_rows：----", excel.get_sheet_rows(sheet))
    print("get_sheet_cols：----", excel.get_sheet_cols(sheet))
    print("get_sheet_data：----", excel.get_sheet_data('百度搜索'))
    # excel.set_column_width(sheet) # 自适应列宽
    # excel.set_row_style(sheet) # 隔行变色
    excel.set_cell_value(sheet,3,8,'你好啊python')
    # excel.set_row_height(sheet,30)  # 设置行高
    # excel.set_cell_style(sheet,5,8,CELL_STYLE_GRAY) # 设置单元格背景颜色
    excel.save(path)  # 保存
